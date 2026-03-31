#!/usr/bin/env python3
"""
Generate Colombia_Teams.xlsx — Lionlive guild team performance tracker.

Usage:
    python3 generate_colombia_tracker.py

Outputs Colombia_Teams.xlsx in the current directory. Re-run to regenerate.

After initial generation:
  - Edit Monthly_Data and Staff sheets directly in Excel — all other sheets
    recalculate via formulas.
  - To activate a new team: change its Status in Config from "Not Launched Yet"
    to "准备中" or "Active". No script needed.
  - To permanently add a new team beyond 10: update TEAMS list and re-run.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE = "Colombia_Teams.xlsx"

# ── Colour palette ─────────────────────────────────────────────────────────────
FILL_HEADER   = PatternFill("solid", fgColor="1F3864")
FILL_SECTION  = PatternFill("solid", fgColor="2E75B6")
FILL_LABEL    = PatternFill("solid", fgColor="D6E4F7")
FILL_ENTRY    = PatternFill("solid", fgColor="FFF2CC")   # yellow  — data entry
FILL_OVERRIDE = PatternFill("solid", fgColor="DDEEFF")   # blue    — actual override
FILL_CALC     = PatternFill("solid", fgColor="F2F2F2")   # grey    — auto-calc
FILL_INACTIVE = PatternFill("solid", fgColor="EBEBEB")   # dimmed  — not launched
FILL_WHITE    = PatternFill("solid", fgColor="FFFFFF")

FONT_HDR  = Font(name="Arial", bold=True,  color="FFFFFF", size=10)
FONT_BOLD = Font(name="Arial", bold=True,  size=10)
FONT_NORM = Font(name="Arial",             size=10)
FONT_DIM  = Font(name="Arial",             size=10, color="AAAAAA")
FONT_TINY = Font(name="Arial",             size=9,  color="666666")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")

THIN = Side(style="thin", color="CCCCCC")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Status vocabulary ──────────────────────────────────────────────────────────
STATUS_NOT_LAUNCHED = "Not Launched Yet"
STATUS_PREPARING    = "准备中"
STATUS_ACTIVE       = "Active"
STATUS_CLOSED       = "Closed"
ALL_STATUSES        = f'"{STATUS_NOT_LAUNCHED},{STATUS_PREPARING},{STATUS_ACTIVE},{STATUS_CLOSED}"'

# ── Team definitions (10 pre-allocated slots) ──────────────────────────────────
# To activate a new team: change status in Config sheet — no script needed.
# To permanently add beyond 10: extend this list and re-run.
TEAMS = [
    {"id": "Lion01", "name": "Tropi",  "market": "Colombia", "launch": "2025-09-01", "status": STATUS_ACTIVE,       "initial_invest": 11430},
    {"id": "Lion02", "name": "Kangri", "market": "Colombia", "launch": "2026-02-01", "status": STATUS_ACTIVE,       "initial_invest": 11430},
    {"id": "Lion03", "name": "",       "market": "Colombia", "launch": "",           "status": STATUS_NOT_LAUNCHED, "initial_invest": 11430},
    {"id": "Lion04", "name": "",       "market": "Colombia", "launch": "",           "status": STATUS_NOT_LAUNCHED, "initial_invest": 11430},
    {"id": "Lion05", "name": "",       "market": "Colombia", "launch": "",           "status": STATUS_NOT_LAUNCHED, "initial_invest": 11430},
    {"id": "Lion06", "name": "",       "market": "Colombia", "launch": "",           "status": STATUS_NOT_LAUNCHED, "initial_invest": 11430},
    {"id": "Lion07", "name": "",       "market": "Colombia", "launch": "",           "status": STATUS_NOT_LAUNCHED, "initial_invest": 11430},
    {"id": "Lion08", "name": "",       "market": "Colombia", "launch": "",           "status": STATUS_NOT_LAUNCHED, "initial_invest": 11430},
    {"id": "Lion09", "name": "",       "market": "Colombia", "launch": "",           "status": STATUS_NOT_LAUNCHED, "initial_invest": 11430},
    {"id": "Lion10", "name": "",       "market": "Colombia", "launch": "",           "status": STATUS_NOT_LAUNCHED, "initial_invest": 11430},
]

# ── Staff definitions ──────────────────────────────────────────────────────────
STAFF = [
    {"name": "Ops Manager",   "role": "Operations",   "salary": 800, "teams": ["Lion01", "Lion02"]},
    {"name": "Dance Teacher", "role": "Dance Teacher", "salary": 600, "teams": ["Lion01"]},
    {"name": "MUA",           "role": "MUA",           "salary": 500, "teams": ["Lion01", "Lion02"]},
    {"name": "Tech Support",  "role": "Tech",          "salary": 400, "teams": ["Lion01", "Lion02"]},
]

# ── Months to include ──────────────────────────────────────────────────────────
MONTHS = ["2026-02", "2026-03"]

# ── Sample data ────────────────────────────────────────────────────────────────
# col_offset: 0=Diamonds, 1=Creator count; 2-6=overrides (blank → Config default)
SAMPLE_DATA = {
    ("2026-02", 0): {0: 890_000,   1: 4},
    ("2026-02", 1): {0: 450_000,   1: 3},
    ("2026-03", 0): {0: 1_200_000, 1: 4},
    ("2026-03", 1): {0: 620_000,   1: 3},
}

# ── Layout constants ───────────────────────────────────────────────────────────
TEAM_BLOCK_COLS = 7
MONTH_COL       = 1

# Config team registry layout (rows)
CONFIG_TEAM_HDR_ROW   = 4
CONFIG_TEAM_START_ROW = 5   # row of Lion01; Lion0N = 5 + (N-1)

# Config team registry columns
CFG_COL_ID      = 1  # A
CFG_COL_NAME    = 2  # B
CFG_COL_MARKET  = 3  # C
CFG_COL_LAUNCH  = 4  # D
CFG_COL_STATUS  = 5  # E
CFG_COL_INVEST  = 6  # F


def team_col_start(team_index: int) -> int:
    """1-based column index of the first column of team block in Monthly_Data."""
    return MONTH_COL + 1 + team_index * TEAM_BLOCK_COLS


def config_row(team_index: int) -> int:
    """Row in Config sheet for this team (0-based index)."""
    return CONFIG_TEAM_START_ROW + team_index


# ── Helpers ────────────────────────────────────────────────────────────────────

def cell(ws, row, col, value=None, fill=None, font=None, align=None,
         fmt=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill   is not None: c.fill          = fill
    if font   is not None: c.font          = font
    if align  is not None: c.alignment     = align
    if fmt    is not None: c.number_format = fmt
    if border is not None: c.border        = border
    return c


def section_hdr(ws, row, col, text, span=2):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = FILL_SECTION; c.font = FONT_HDR; c.alignment = ALIGN_CENTER
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return c


def title_row(ws, row, text, total_cols, height=28):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=total_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = FILL_HEADER
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    c.alignment = ALIGN_CENTER
    ws.row_dimensions[row].height = height
    return c


def add_named_range(wb, name, sheet_name, cell_ref):
    wb.defined_names[name] = DefinedName(
        name, attr_text=f"'{sheet_name}'!{cell_ref}"
    )


def wrap_if_active(status_ref, formula):
    """
    Wrap an Excel formula so it returns "" when team is Not Launched Yet.
    formula must start with '='.
    Active, 准备中, and Closed all display their data.
    """
    inner = formula[1:]   # strip leading '='
    return f'=IF({status_ref}="{STATUS_NOT_LAUNCHED}","",{inner})'


# ── Sheet builders ─────────────────────────────────────────────────────────────

def build_config(wb):
    ws = wb.create_sheet("Config")
    n_teams = len(TEAMS)

    # Column widths
    for col, w in zip("ABCDEF", [14, 18, 14, 14, 18, 18]):
        ws.column_dimensions[col].width = w
    ws.column_dimensions["G"].width = 28

    title_row(ws, 1, "⚙️  CONFIG — Edit here to manage teams and assumptions", 7)

    # ── Section A: Team Registry ───────────────────────────────────
    section_hdr(ws, 3, 1,
        "TEAM REGISTRY  ←  Change Status to activate/close teams (no script needed)",
        span=7)

    for ci, hdr in enumerate(
        ["Team ID", "Team Name", "Market", "Launch Date", "Status", "Initial Invest ($)"],
        1
    ):
        cell(ws, CONFIG_TEAM_HDR_ROW, ci, hdr,
             fill=FILL_LABEL, font=FONT_BOLD, align=ALIGN_CENTER, border=BORDER_THIN)

    # Status dropdown validation
    dv = DataValidation(
        type="list",
        formula1=ALL_STATUSES,
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid status",
        error="Choose: Not Launched Yet / 准备中 / Active / Closed"
    )
    ws.add_data_validation(dv)

    for ti, team in enumerate(TEAMS):
        row = config_row(ti)
        is_active = team["status"] != STATUS_NOT_LAUNCHED
        row_fill  = FILL_WHITE if is_active else FILL_INACTIVE
        row_font  = FONT_NORM  if is_active else FONT_DIM

        cell(ws, row, CFG_COL_ID,     team["id"],             fill=FILL_LABEL,   font=FONT_BOLD, align=ALIGN_CENTER, border=BORDER_THIN)
        cell(ws, row, CFG_COL_NAME,   team["name"],           fill=FILL_ENTRY if is_active else FILL_INACTIVE, font=row_font, align=ALIGN_CENTER, border=BORDER_THIN)
        cell(ws, row, CFG_COL_MARKET, team["market"],         fill=FILL_ENTRY if is_active else FILL_INACTIVE, font=row_font, align=ALIGN_CENTER, border=BORDER_THIN)
        cell(ws, row, CFG_COL_LAUNCH, team["launch"] or None, fill=FILL_ENTRY if is_active else FILL_INACTIVE, font=row_font, align=ALIGN_CENTER, fmt="YYYY-MM-DD", border=BORDER_THIN)
        cell(ws, row, CFG_COL_STATUS, team["status"],         fill=FILL_ENTRY, font=FONT_BOLD if is_active else FONT_DIM, align=ALIGN_CENTER, border=BORDER_THIN)
        cell(ws, row, CFG_COL_INVEST, team["initial_invest"], fill=FILL_ENTRY, font=FONT_NORM, align=ALIGN_CENTER, fmt="$#,##0", border=BORDER_THIN)

        # Hint column
        hint = "" if is_active else "← Change Status to activate"
        cell(ws, row, 7, hint, font=FONT_TINY, align=ALIGN_LEFT)

        # Apply status dropdown to status column
        dv.add(ws.cell(row=row, column=CFG_COL_STATUS))

    # Status key
    key_row = config_row(n_teams) + 1
    cell(ws, key_row, 1, "Status key:", font=FONT_BOLD)
    for ci, (status, desc) in enumerate([
        (STATUS_NOT_LAUNCHED, "Slot reserved, not started"),
        (STATUS_PREPARING,    "Pre-launch costs accumulating, no revenue"),
        (STATUS_ACTIVE,       "Fully operational"),
        (STATUS_CLOSED,       "Historical data preserved"),
    ], 2):
        cell(ws, key_row, ci, f"{status} = {desc}", font=FONT_TINY)
        key_row += 1

    # ── Section B: UE Assumptions ──────────────────────────────────
    row = config_row(n_teams) + 7
    section_hdr(ws, row, 1,
                "UE ASSUMPTIONS  (edit yellow cells — apply to all teams unless overridden)",
                span=3)
    row += 1

    assumptions = [
        ("Take Rate",              0.65, "0%",      "CONFIG_TAKE_RATE"),
        ("Diamond → USD rate",     0.01, "$0.0000", "CONFIG_DIAMOND_RATE"),
        ("Monthly Rent (USD)",     600,  "$#,##0",  "CONFIG_RENT"),
        ("Monthly Utility (USD)",  800,  "$#,##0",  "CONFIG_UTILITY"),
        ("Monthly Buffer (USD)",   1000, "$#,##0",  "CONFIG_BUFFER"),
        ("Default Initial Invest", 11430,"$#,##0",  "CONFIG_DEFAULT_INITIAL_INVEST"),
        ("Creator Base Salary",    500,  "$#,##0",  "CONFIG_CREATOR_BASE"),
        ("Revenue Share %",        0.25, "0%",      "CONFIG_CREATOR_REVSHARE"),
    ]
    for label, val, fmt, name in assumptions:
        cell(ws, row, 1, label, fill=FILL_LABEL, font=FONT_BOLD, align=ALIGN_LEFT, border=BORDER_THIN)
        cell(ws, row, 2, val,   fill=FILL_ENTRY, font=FONT_NORM, align=ALIGN_CENTER, fmt=fmt, border=BORDER_THIN)
        cell(ws, row, 3, f"← Named: {name}", font=FONT_TINY)
        add_named_range(wb, name, "Config", f"$B${row}")
        row += 1

    # ── Section C: Stage Thresholds ───────────────────────────────
    row += 1
    section_hdr(ws, row, 1, "STAGE THRESHOLDS (diamonds / month)", span=3)
    row += 1
    for ci, hdr in enumerate(["Stage", "Min Diamonds (≥)"], 1):
        cell(ws, row, ci, hdr, fill=FILL_LABEL, font=FONT_BOLD, align=ALIGN_CENTER, border=BORDER_THIN)
    row += 1

    for stage_label, stage_val, *name_opt in [
        ("🌱 Incubation",  0),
        ("⚠️ Breakeven",   500_000,   "CONFIG_STAGE_BE"),
        ("✅ Stable",      1_060_000,  "CONFIG_STAGE_STB"),
        ("🚀 Optimistic",  2_000_000,  "CONFIG_STAGE_OPT"),
    ]:
        cell(ws, row, 1, stage_label, font=FONT_NORM, border=BORDER_THIN)
        cell(ws, row, 2, stage_val,   fill=FILL_ENTRY, font=FONT_NORM, align=ALIGN_CENTER, fmt="#,##0", border=BORDER_THIN)
        if name_opt:
            add_named_range(wb, name_opt[0], "Config", f"$B${row}")
        row += 1

    # ── Legend ─────────────────────────────────────────────────────
    row += 1
    section_hdr(ws, row, 1, "LEGEND", span=2)
    row += 1
    for fill, label in [
        (FILL_ENTRY,    "Yellow = Data entry required"),
        (FILL_OVERRIDE, "Blue   = Actual override (blank → use Config assumption)"),
        (FILL_CALC,     "Grey   = Auto-calculated — do not edit"),
        (FILL_INACTIVE, "Light grey = Team slot not yet activated"),
    ]:
        ws.cell(row=row, column=1).fill = fill
        cell(ws, row, 2, label, font=FONT_TINY)
        row += 1


def build_staff(wb):
    ws = wb.create_sheet("Staff")
    team_ids = [t["id"] for t in TEAMS]
    n_teams  = len(team_ids)

    col_first_chk = 4
    col_covered   = col_first_chk + n_teams
    col_coeff     = col_covered + 1
    col_cost_s    = col_coeff + 1
    total_cols    = col_cost_s + n_teams - 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    for i in range(n_teams):
        ws.column_dimensions[get_column_letter(col_first_chk + i)].width = 9
    ws.column_dimensions[get_column_letter(col_covered)].width = 14
    ws.column_dimensions[get_column_letter(col_coeff)].width   = 14
    for i in range(n_teams):
        ws.column_dimensions[get_column_letter(col_cost_s + i)].width = 14

    title_row(ws, 1,
        '👥  STAFF — Mark "✓" to assign teams; allocation & cost auto-calculate',
        total_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    cell(ws, 2, 1,
         "Alloc Coeff = 1 ÷ Teams Covered.  Per-team cost = Salary × Coeff (if assigned).  "
         "Add new staff rows freely — SUM range covers up to row 200.",
         font=Font(name="Arial", italic=True, size=9, color="444444"),
         align=ALIGN_LEFT)

    # Headers row 3
    for ci, hdr in enumerate(["Staff Name", "Role", "Salary (USD)"], 1):
        cell(ws, 3, ci, hdr, fill=FILL_LABEL, font=FONT_BOLD, align=ALIGN_CENTER, border=BORDER_THIN)
    for i, tid in enumerate(team_ids):
        team_active = TEAMS[i]["status"] != STATUS_NOT_LAUNCHED
        cell(ws, 3, col_first_chk + i, tid,
             fill=FILL_LABEL if team_active else FILL_INACTIVE,
             font=FONT_BOLD if team_active else FONT_DIM,
             align=ALIGN_CENTER, border=BORDER_THIN)
    cell(ws, 3, col_covered, "Teams\nCovered",
         fill=FILL_CALC, font=FONT_BOLD, align=ALIGN_CENTER, border=BORDER_THIN)
    cell(ws, 3, col_coeff,   "Alloc\nCoeff",
         fill=FILL_CALC, font=FONT_BOLD, align=ALIGN_CENTER, border=BORDER_THIN)
    for i, tid in enumerate(team_ids):
        team_active = TEAMS[i]["status"] != STATUS_NOT_LAUNCHED
        cell(ws, 3, col_cost_s + i, f"{tid}\nCost",
             fill=FILL_CALC if team_active else FILL_INACTIVE,
             font=FONT_BOLD if team_active else FONT_DIM,
             align=ALIGN_CENTER, border=BORDER_THIN)

    # Staff data rows
    for r_off, staff in enumerate(STAFF):
        row = 4 + r_off
        cell(ws, row, 1, staff["name"], font=FONT_NORM, border=BORDER_THIN)
        cell(ws, row, 2, staff["role"],  font=FONT_NORM, border=BORDER_THIN)
        cell(ws, row, 3, staff["salary"],
             fill=FILL_ENTRY, font=FONT_NORM, align=ALIGN_RIGHT, fmt="$#,##0", border=BORDER_THIN)

        chk_letters = []
        for i, tid in enumerate(team_ids):
            col = col_first_chk + i
            val = "✓" if tid in staff["teams"] else ""
            team_active = TEAMS[i]["status"] != STATUS_NOT_LAUNCHED
            cell(ws, row, col, val,
                 fill=FILL_ENTRY if team_active else FILL_INACTIVE,
                 font=FONT_NORM if team_active else FONT_DIM,
                 align=ALIGN_CENTER, border=BORDER_THIN)
            chk_letters.append(get_column_letter(col))

        cov_letter = get_column_letter(col_covered)
        cell(ws, row, col_covered,
             f'=COUNTIF({chk_letters[0]}{row}:{chk_letters[-1]}{row},"✓")',
             fill=FILL_CALC, font=FONT_NORM, align=ALIGN_CENTER, fmt="0", border=BORDER_THIN)
        cell(ws, row, col_coeff,
             f"=IF({cov_letter}{row}=0,0,1/{cov_letter}{row})",
             fill=FILL_CALC, font=FONT_NORM, align=ALIGN_CENTER, fmt="0.00", border=BORDER_THIN)

        coeff_l = get_column_letter(col_coeff)
        sal_l   = get_column_letter(3)
        for i, tid in enumerate(team_ids):
            chk_l = chk_letters[i]
            cell(ws, row, col_cost_s + i,
                 f'=IF({chk_l}{row}="✓",{sal_l}{row}*{coeff_l}{row},0)',
                 fill=FILL_CALC, font=FONT_NORM, align=ALIGN_RIGHT, fmt="$#,##0.00", border=BORDER_THIN)

    # Totals row
    tot_row = 4 + len(STAFF)
    cell(ws, tot_row, 1, "TOTAL", font=FONT_BOLD, fill=FILL_LABEL, border=BORDER_THIN)
    cell(ws, tot_row, 3, f"=SUM(C4:C{tot_row-1})",
         fill=FILL_LABEL, font=FONT_BOLD, fmt="$#,##0", border=BORDER_THIN)
    for i in range(n_teams):
        cl = get_column_letter(col_cost_s + i)
        cell(ws, tot_row, col_cost_s + i, f"=SUM({cl}4:{cl}{tot_row-1})",
             fill=FILL_LABEL, font=FONT_BOLD, fmt="$#,##0.00", border=BORDER_THIN)

    ws._col_cost_start  = col_cost_s
    ws._staff_data_rows = (4, 200)   # generous range: manually added rows auto-included


def build_monthly_data(wb):
    ws = wb.create_sheet("Monthly_Data")
    team_ids = [t["id"] for t in TEAMS]
    n_teams  = len(team_ids)

    admin_col_s = team_col_start(n_teams)
    last_col    = admin_col_s + 3

    ws.column_dimensions["A"].width = 12

    title_row(ws, 1,
        "📋  MONTHLY DATA — Yellow = required. Blue = actual override (blank → Config default).",
        last_col)

    # Team block headers (row 2)
    for ti, tid in enumerate(team_ids):
        col_s = team_col_start(ti)
        ws.merge_cells(start_row=2, start_column=col_s,
                       end_row=2, end_column=col_s + TEAM_BLOCK_COLS - 1)
        team_active = TEAMS[ti]["status"] != STATUS_NOT_LAUNCHED
        cell(ws, 2, col_s, f"── {tid} ──",
             fill=FILL_SECTION if team_active else FILL_INACTIVE,
             font=FONT_HDR if team_active else FONT_DIM,
             align=ALIGN_CENTER)

    ws.merge_cells(start_row=2, start_column=admin_col_s,
                   end_row=2, end_column=admin_col_s + 3)
    cell(ws, 2, admin_col_s, "── SHARED ADMIN (split equally across Active teams) ──",
         fill=FILL_SECTION, font=FONT_HDR, align=ALIGN_CENTER)

    # Column sub-headers (row 3)
    cell(ws, 3, 1, "Month",
         fill=FILL_LABEL, font=FONT_BOLD, align=ALIGN_CENTER, border=BORDER_THIN)

    team_sub = [
        ("Diamonds",            FILL_ENTRY,    "#,##0"),
        ("Creators",            FILL_ENTRY,    "0"),
        ("Creator Sal. Actual", FILL_OVERRIDE, "$#,##0"),
        ("Rent Actual",         FILL_OVERRIDE, "$#,##0"),
        ("Utility Actual",      FILL_OVERRIDE, "$#,##0"),
        ("Buffer Actual",       FILL_OVERRIDE, "$#,##0"),
        ("Other Cost Actual",   FILL_OVERRIDE, "$#,##0"),
    ]
    for ti in range(n_teams):
        col_s       = team_col_start(ti)
        team_active = TEAMS[ti]["status"] != STATUS_NOT_LAUNCHED
        for ci, (label, fill, _fmt) in enumerate(team_sub):
            col = col_s + ci
            c_fill = fill if team_active else FILL_INACTIVE
            cell(ws, 3, col, label,
                 fill=c_fill, font=FONT_BOLD if team_active else FONT_DIM,
                 align=ALIGN_CENTER, border=BORDER_THIN)
            ws.column_dimensions[get_column_letter(col)].width = 13

    for ci, label in enumerate(["Clothes", "Taxi", "Meals", "Other Admin"]):
        col = admin_col_s + ci
        cell(ws, 3, col, label,
             fill=FILL_ENTRY, font=FONT_BOLD, align=ALIGN_CENTER, border=BORDER_THIN)
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Data rows
    data_row_start = 4
    for ri, month in enumerate(MONTHS):
        row = data_row_start + ri
        cell(ws, row, 1, month,
             fill=FILL_WHITE, font=FONT_NORM, align=ALIGN_CENTER, border=BORDER_THIN)

        for ti in range(n_teams):
            col_s       = team_col_start(ti)
            team_active = TEAMS[ti]["status"] != STATUS_NOT_LAUNCHED
            for ci, (_label, fill, fmt) in enumerate(team_sub):
                col = col_s + ci
                val = None
                key = (month, ti)
                if key in SAMPLE_DATA and ci in SAMPLE_DATA[key]:
                    val = SAMPLE_DATA[key][ci]
                cell(ws, row, col, val,
                     fill=fill if team_active else FILL_INACTIVE,
                     font=FONT_NORM if team_active else FONT_DIM,
                     fmt=fmt, border=BORDER_THIN)

        for ci in range(4):
            col = admin_col_s + ci
            cell(ws, row, col, None,
                 fill=FILL_ENTRY, font=FONT_NORM, fmt="$#,##0", border=BORDER_THIN)

    ws.freeze_panes = "B4"

    ws._admin_col_start = admin_col_s
    ws._data_row_start  = data_row_start
    ws._data_row_end    = data_row_start + len(MONTHS) - 1


def build_ue_summary(wb):
    ws     = wb.create_sheet("UE_Summary")
    md_ws  = wb["Monthly_Data"]
    st_ws  = wb["Staff"]
    n_teams = len(TEAMS)

    # Count active teams for admin split (use formula-based count for live updates)
    # We'll use COUNTIF on Config status column
    cfg_status_range = (
        f"'Config'!${get_column_letter(CFG_COL_STATUS)}"
        f"${CONFIG_TEAM_START_ROW}:"
        f"${get_column_letter(CFG_COL_STATUS)}"
        f"${CONFIG_TEAM_START_ROW + n_teams - 1}"
    )
    # Admin split denominator: count of Active + 准备中 + Closed teams
    admin_denom_f = (
        f'COUNTIF({cfg_status_range},"{STATUS_ACTIVE}")'
        f'+COUNTIF({cfg_status_range},"{STATUS_PREPARING}")'
        f'+COUNTIF({cfg_status_range},"{STATUS_CLOSED}")'
    )

    headers = [
        "Month", "Team ID", "Team Name",
        "Diamonds", "Revenue (USD)",
        "Creator Salary", "Staff Cost",
        "Rent", "Utility", "Buffer", "Admin Share", "Other",
        "Total Cost", "Gross Profit", "Margin %",
        "Stage",
        "Cumul. Net", "Initial Invest.", "Recovered?", "Month #"
    ]
    fmts = [
        None, None, None,
        "#,##0", "$#,##0.00",
        "$#,##0.00", "$#,##0.00",
        "$#,##0.00", "$#,##0.00", "$#,##0.00", "$#,##0.00", "$#,##0.00",
        "$#,##0.00", "$#,##0.00", "0.0%",
        None,
        "$#,##0.00", "$#,##0.00", None, "0"
    ]

    title_row(ws, 1, "📊  UE SUMMARY — Auto-calculated. Do not edit.", len(headers))

    for ci, (hdr, _) in enumerate(zip(headers, fmts), 1):
        cell(ws, 2, ci, hdr,
             fill=FILL_CALC, font=FONT_BOLD, align=ALIGN_CENTER, border=BORDER_THIN)
        ws.column_dimensions[get_column_letter(ci)].width = 14

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16
    ws.freeze_panes = "D3"

    team_profit_rows = {t["id"]: [] for t in TEAMS}

    data_row = 3
    for month_idx, month in enumerate(MONTHS):
        md_row = md_ws._data_row_start + month_idx

        for ti, team in enumerate(TEAMS):
            tid  = team["id"]
            name = team["name"]
            cs   = team_col_start(ti)

            # Config references for this team
            cfg_r        = config_row(ti)
            status_ref   = f"'Config'!${get_column_letter(CFG_COL_STATUS)}${cfg_r}"
            invest_ref   = f"'Config'!${get_column_letter(CFG_COL_INVEST)}${cfg_r}"
            launch_ref   = f"'Config'!${get_column_letter(CFG_COL_LAUNCH)}${cfg_r}"

            # Monthly_Data column letters
            dia_l = get_column_letter(cs)
            crt_l = get_column_letter(cs + 1)
            csa_l = get_column_letter(cs + 2)
            rnt_l = get_column_letter(cs + 3)
            utl_l = get_column_letter(cs + 4)
            buf_l = get_column_letter(cs + 5)
            oth_l = get_column_letter(cs + 6)

            dia_ref = f"Monthly_Data!{dia_l}{md_row}"
            crt_ref = f"Monthly_Data!{crt_l}{md_row}"
            csa_ref = f"Monthly_Data!{csa_l}{md_row}"

            adm_start_l = get_column_letter(md_ws._admin_col_start)
            adm_end_l   = get_column_letter(md_ws._admin_col_start + 3)
            admin_range = f"Monthly_Data!{adm_start_l}{md_row}:{adm_end_l}{md_row}"

            # Core expressions (no leading =)
            rev_expr = f"{dia_ref}*CONFIG_DIAMOND_RATE*CONFIG_TAKE_RATE"

            diamonds_f = f"=IF({dia_ref}=\"\",0,{dia_ref})"
            revenue_f  = f"=IF({dia_ref}=\"\",0,{rev_expr})"

            creator_f  = (
                f"=IF({csa_ref}<>\"\",{csa_ref},"
                f"MAX(CONFIG_CREATOR_BASE*IF({crt_ref}=\"\",0,{crt_ref}),"
                f"{rev_expr}*CONFIG_CREATOR_REVSHARE))"
            )

            sc_col  = get_column_letter(st_ws._col_cost_start + ti)
            sr_s, sr_e = st_ws._staff_data_rows
            staff_f = f"=SUM(Staff!{sc_col}{sr_s}:Staff!{sc_col}{sr_e})"

            def override_f(actual_l, config_name):
                return (f"=IF(Monthly_Data!{actual_l}{md_row}<>\"\","
                        f"Monthly_Data!{actual_l}{md_row},{config_name})")

            rent_f    = override_f(rnt_l, "CONFIG_RENT")
            utility_f = override_f(utl_l, "CONFIG_UTILITY")
            buffer_f  = override_f(buf_l, "CONFIG_BUFFER")
            other_f   = f"=IF(Monthly_Data!{oth_l}{md_row}<>\"\",Monthly_Data!{oth_l}{md_row},0)"
            admin_f   = f"=IF(({admin_denom_f})=0,0,SUM({admin_range})/({admin_denom_f}))"

            total_f  = f"=SUM(F{data_row}:L{data_row})"
            profit_f = f"=E{data_row}-M{data_row}"
            margin_f = f"=IF(E{data_row}=0,0,N{data_row}/E{data_row})"

            stage_f = (
                f'=IFS(D{data_row}>=CONFIG_STAGE_OPT,"🚀 Optimistic",'
                f'D{data_row}>=CONFIG_STAGE_STB,"✅ Stable",'
                f'D{data_row}>=CONFIG_STAGE_BE,"⚠️ Breakeven",'
                f'TRUE,"🌱 Incubation")'
            )

            prior_rows = team_profit_rows[tid]
            if prior_rows:
                prior_sum = "+".join(f"N{r}" for r in prior_rows)
                cumul_f = f"=N{data_row}+{prior_sum}"
            else:
                cumul_f = f"=N{data_row}"

            # Per-team initial investment from Config (not a global named range)
            invest_f = f"={invest_ref}"

            recovered_f = f'=IF(Q{data_row}>={invest_ref},"✅ YES","⏳ No")'

            # Month # from Config launch date (handles blank launch date gracefully)
            month_num_f = (
                f'=IF({launch_ref}="","—",'
                f'DATEDIF(DATE(LEFT(TEXT({launch_ref},"YYYY-MM-DD"),4),'
                f'MID(TEXT({launch_ref},"YYYY-MM-DD"),6,2),1),'
                f'DATE(LEFT(A{data_row},4),RIGHT(A{data_row},2),1),"M")+1)'
            )

            row_values = [
                (month,      None),
                (tid,        None),
                (name or f"({tid})", None),
                (diamonds_f, "#,##0"),
                (revenue_f,  "$#,##0.00"),
                (creator_f,  "$#,##0.00"),
                (staff_f,    "$#,##0.00"),
                (rent_f,     "$#,##0.00"),
                (utility_f,  "$#,##0.00"),
                (buffer_f,   "$#,##0.00"),
                (admin_f,    "$#,##0.00"),
                (other_f,    "$#,##0.00"),
                (total_f,    "$#,##0.00"),
                (profit_f,   "$#,##0.00"),
                (margin_f,   "0.0%"),
                (stage_f,    None),
                (cumul_f,    "$#,##0.00"),
                (invest_f,   "$#,##0.00"),
                (recovered_f,None),
                (month_num_f,"0"),
            ]

            is_active = team["status"] != STATUS_NOT_LAUNCHED

            for ci, (val, fmt) in enumerate(row_values, 1):
                # Wrap calculated formulas (cols 4–20) with NOT LAUNCHED guard
                if isinstance(val, str) and val.startswith("=") and ci > 3:
                    val = wrap_if_active(status_ref, val)
                c = ws.cell(row=data_row, column=ci, value=val)
                c.fill   = (FILL_CALC  if ci > 3 else FILL_WHITE) if is_active else FILL_INACTIVE
                c.font   = FONT_NORM if is_active else FONT_DIM
                c.border = BORDER_THIN
                c.alignment = ALIGN_CENTER
                if fmt:
                    c.number_format = fmt

            team_profit_rows[tid].append(data_row)
            data_row += 1

    ws._data_row_start    = 3
    ws._data_row_end      = data_row - 1
    ws._team_profit_rows  = team_profit_rows


def build_team_sheet(wb, team, ti):
    tid  = team["id"]
    name = team["name"] or f"({tid})"
    ws   = wb.create_sheet(f"Team_{tid}")
    ue   = wb["UE_Summary"]

    team_ue_rows = ue._team_profit_rows[tid]
    is_active    = team["status"] != STATUS_NOT_LAUNCHED
    status_ref   = f"'Config'!${get_column_letter(CFG_COL_STATUS)}${config_row(ti)}"

    # Title
    title_row(ws, 1, f"🏠  {tid} — {name}  ({team['market']})", 18)

    if not is_active:
        ws.merge_cells("A2:R2")
        cell(ws, 2, 1,
             f"Status: {team['status']} — Change status in Config to activate this team.",
             font=FONT_DIM, align=ALIGN_CENTER)
        return

    # Meta info
    for ci, (label, val) in enumerate([
        ("Launch Date:", team["launch"] or "—"),
        ("Status:", team["status"]),
        ("Market:", team["market"]),
    ]):
        cell(ws, 2, 1 + ci * 3,     label, font=FONT_BOLD)
        cell(ws, 2, 2 + ci * 3,     val,   font=FONT_NORM)

    # Monthly table headers
    col_headers = [
        "Month", "Diamonds", "Revenue", "Creator Sal.", "Staff Cost",
        "Rent", "Utility", "Buffer", "Admin", "Other",
        "Total Cost", "Gross Profit", "Margin %",
        "Stage", "Cumul. Net", "Initial Invest.", "Recovered?", "Month #"
    ]
    # UE_Summary columns: A=1, D=4, E=5 … T=20
    ue_col_map = [1, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]
    col_fmts = {
        2: "#,##0", 3: "$#,##0.00", 4: "$#,##0.00",  5: "$#,##0.00",
        6: "$#,##0.00", 7: "$#,##0.00", 8: "$#,##0.00", 9: "$#,##0.00",
        10: "$#,##0.00", 11: "$#,##0.00", 12: "$#,##0.00", 13: "0.0%",
        15: "$#,##0.00", 16: "$#,##0.00", 18: "0"
    }

    hdr_row = 4
    for ci, hdr in enumerate(col_headers, 1):
        cell(ws, hdr_row, ci, hdr,
             fill=FILL_LABEL, font=FONT_BOLD, align=ALIGN_CENTER, border=BORDER_THIN)
        ws.column_dimensions[get_column_letter(ci)].width = 13

    for ri, ue_row in enumerate(team_ue_rows):
        row = hdr_row + 1 + ri
        for ci, ue_col in enumerate(ue_col_map, 1):
            ue_col_l = get_column_letter(ue_col)
            c = ws.cell(row=row, column=ci,
                        value=f"=UE_Summary!{ue_col_l}{ue_row}")
            c.fill = FILL_CALC; c.font = FONT_NORM
            c.border = BORDER_THIN; c.alignment = ALIGN_CENTER
            if ci in col_fmts:
                c.number_format = col_fmts[ci]

    # Investment recovery line chart
    if team_ue_rows:
        data_start = hdr_row + 1
        data_end   = hdr_row + len(team_ue_rows)
        chart_row  = data_end + 3

        chart = LineChart()
        chart.title  = f"{tid} — Cumulative Net Profit vs Break-even"
        chart.style  = 10
        chart.height = 12
        chart.width  = 22
        chart.y_axis.title = "USD"
        chart.x_axis.title = "Month"

        data_ref = Reference(ws, min_col=15, min_row=data_start, max_row=data_end)
        chart.add_data(data_ref, titles_from_data=False)
        chart.series[0].title = SeriesLabel(v="Cumulative Net Profit")
        chart.series[0].graphicalProperties.line.solidFill = "2E75B6"
        chart.series[0].graphicalProperties.line.width = 20000

        cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{chart_row}")


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_config(wb)
    build_staff(wb)
    build_monthly_data(wb)
    build_ue_summary(wb)
    for ti, team in enumerate(TEAMS):
        build_team_sheet(wb, team, ti)

    wb.save(OUTPUT_FILE)
    active = sum(1 for t in TEAMS if t["status"] == STATUS_ACTIVE)
    print(f"✅  Saved {OUTPUT_FILE}")
    print(f"    Sheets : {', '.join(ws.title for ws in wb.worksheets)}")
    print(f"    Teams  : {len(TEAMS)} slots ({active} Active) | Months: {len(MONTHS)} | Staff: {len(STAFF)}")


if __name__ == "__main__":
    main()
